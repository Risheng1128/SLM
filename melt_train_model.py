import xgboost as xgb
import lightgbm as gbm
import pandas as pd
import numpy as np
import openpyxl
import argparse
import pickle
import os
import melt_constants as const

from sklearn import metrics
from sklearn import svm
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import LabelEncoder
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import LogisticRegression
from sklearn.feature_selection import mutual_info_regression
from sklearn.feature_selection import SelectKBest
from sklearn.decomposition import PCA
from sklearn.manifold import TSNE
from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import RandomizedSearchCV

class data:
    def __init__(self, keys=None, datas=None):
        if datas is None:
            self.__data = {key: [] for key in keys}
        elif isinstance(datas, dict):
            self.__data = datas
        elif isinstance(datas, list):
            self.__data = {key: data for key, data in zip(keys, datas)}
        else:
            self.__data = {key: datas for key in keys}

    def append(self, key, data):
        self.__data[key].append(data)

    def reshape(self, key, shape):
        self.__data[key] = np.reshape(self.__data[key], shape)

    def repeat(self, key, repeat):
        self.__data[key] = np.repeat(self.__data[key], repeat)

    def unique(self, key, k):
        self.__data[key] = self.__data[key][::k]

    def keys(self):
        return self.__data.keys()

    # return all data
    def read_all(self):
        return [self.__data[key] for key in self.__data.keys()]

    # write data based on the keys of input data
    def write_all(self, datas):
        for key in datas.keys():
            if key in self.__data:
                self.__data[key] = datas[key]

    # read data based on key
    def read(self, key):
        return self.__data[key]

    # write data based on key
    def write(self, key, data):
        self.__data[key] = data

class dataset:
    def __init__(self, keys, output, feature_num=13):
        self.__x_train = data(keys=keys)
        self.__x_test = data(keys=keys)
        self.__y_train = data(keys=keys)
        self.__y_test = data(keys=keys)
        self.__xgboost = data(datas=const.xgboost_param)
        self.__lightgbm = data(datas=const.lightgbm_param)
        self.__logistic = data(datas=const.logistic_param)
        self.__svr = data(datas=const.svr_param)
        self.__feature_num = data(keys, feature_num)
        self.__remove_feature = data(keys=keys)
        self.__keys = keys
        self.__output = output
        self.__use_proc_param = False

    # read train and test data based on key
    def __read_train_and_test(self, key):
        return self.__x_train.read(key), self.__x_test.read(key), \
            self.__y_train.read(key), self.__y_test.read(key)

    # do standard scaler tp x_train and x_test
    def __standard_scaler(self, x_train, x_test):
        ss = StandardScaler()
        x_train = ss.fit_transform(x_train)
        x_test = ss.fit_transform(x_test)
        return x_train, x_test

    # store data into excel (row direction)
    def __store_row_data(self, sheet, datas, base_row, base_col):
        for col, data in zip(range(base_col, base_col + len(datas)), datas):
            sheet.cell(base_row, col).value = data

    # store model result into excel
    def __store_result_data(self, key, sheet, predict):
        x_train, x_test, _, y_test = self.__read_train_and_test(key)
        datas = [x_train.shape[0],
                 x_test.shape[0],
                 self.__feature_num.read(key),
                 ', '.join(map(str, self.__remove_feature.read(key))),
                 metrics.r2_score(y_test, predict),
                 metrics.mean_squared_error(y_test, predict),
                 metrics.mean_absolute_error(y_test, predict)]

        self.__store_row_data(sheet, const.output_label, 1, 1)
        self.__store_row_data(sheet, datas, 2, 4)

        row = 2
        for pre, true in zip(predict, y_test):
            sheet.cell(row, 1).value = pre
            sheet.cell(row, 2).value = true
            sheet.cell(row, 3).value = (abs(pre - true) / true) * 100
            row += 1

    # store model setting into excel
    def __store_model_set(self, sheet, model_set):
        keys = model_set.keys()
        datas = [model_set.read(key) for key in keys]
        self.__store_row_data(sheet, keys, 3, 4)
        self.__store_row_data(sheet, datas, 4, 4)

    # load workpiece and material property data from excel
    # wpfp: workpiece filepath
    # ppfp: property filepath
    def load_data(self, wpfp, ppfp, header=[], use_proc_param=False):
        remove_label = const.layer_label + header

        # wpf: workpiece file
        # ppf: property file
        for wpf, ppf in zip(wpfp, ppfp):
            print('read workpiece file: ', wpf)
            print('read material property file: ', ppf)
            wp_sheets = pd.read_excel(wpf, sheet_name=None)

            # wpd: workpiece data
            # ppd: property data
            for key in self.__keys:
                # the dictionary key must same in property excel
                ppd = pd.read_excel(ppf, sheet_name=key)
                ppd = np.array(ppd.drop(const.trail_label, axis=1))
                # split data
                if use_proc_param:
                    # set totoal feature number
                    self.__feature_num.write(key, 18)
                    self.__use_proc_param = True
                    proc_param = ppd[:, 0:const.printer_param_col]
                ppd = ppd[:, const.printer_param_col::]
                # switch 2D array to 1D array
                ppd = np.reshape(ppd, -1)

                for index, sheet in zip(range(len(ppd)), wp_sheets):
                    trail = int(sheet[5]) - 1
                    item = sheet[-1]
                    # check the property is legal
                    if ppd[index] == const.ignore_data:
                        continue

                    wpd = pd.read_excel(wpf, sheet_name=sheet)
                    wpd = np.array(wpd.drop(remove_label, axis=1))
                    # add process parameter into train and test data
                    if use_proc_param:
                        wpd = np.array([np.concatenate([wp, proc_param[trail]])
                                        for wp in wpd])
                    if item == '5' or item == '6':
                        self.__x_test.append(key, wpd)
                        self.__y_test.append(key, ppd[index])
                    else:
                        self.__x_train.append(key, wpd)
                        self.__y_train.append(key, ppd[index])

    # change the data shape or repeat same data
    def reshape_and_repeat(self, shape, repeat=1):
        for key in self.__keys:
            # convert 3-D matrix to 2-D matrix
            self.__x_train.reshape(key, shape)
            self.__x_test.reshape(key, shape)
            self.__y_train.repeat(key, repeat)
            self.__y_test.repeat(key, repeat)

    # get all unique data in repeated array
    def unique(self, k):
        for key in self.__keys:
            self.__y_train.unique(key, k)
            self.__y_test.unique(key, k)

    # show the correlation coefficient
    def show_corrcoef(self):
        for key in self.keys:
            x_train, x_test, y_train, y_test = self.__read_train_and_test(key)
            x = np.vstack((x_train, x_test))
            y = np.concatenate((y_train, y_test))
            x_mean = np.mean(x, axis=0)
            y_mean = np.mean(y)

            numerator = np.zeros(shape=x_mean.shape[0])
            denominator1 = np.zeros(shape=x_mean.shape[0])
            denominator2 = np.zeros(shape=x_mean.shape[0])

            # compute correlation coefficient
            for xi, yi in zip(x, y):
                x_tmp = xi - x_mean
                y_tmp = yi - y_mean
                numerator += x_tmp * y_tmp
                denominator1 += x_tmp ** 2
                denominator2 += y_tmp ** 2

            r = numerator / (np.sqrt(denominator1) * np.sqrt(denominator2))
            if self.__use_proc_param:
                features = const.feature + const.proc_param
            print('----------', key, '----------')
            for feature, i in zip(features, r):
                print(feature, ': ', i)

    # compute mutual information and retain "k" best data
    def mutual_information(self, k=3):
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            selector = SelectKBest(mutual_info_regression, k=k)
            self.__x_train.write(key, selector.fit_transform(x_train, y_train))
            # find the index of removed feature
            supports = selector.get_support()
            false_index = np.where(supports == False)[0]
            self.__x_test.write(key, np.delete(x_test, false_index, axis=1))

            # record retained feature number
            self.__feature_num.write(key, k)

            if self.__use_proc_param:
                features = const.feature + const.proc_param
            # record removed feature
            for feature, support in zip(features, supports):
                if support == False:
                    self.__remove_feature.append(key, feature)

    # principal component analysis
    def PCA(self, components):
        pca = PCA(n_components=components)
        for key in self.__keys:
            pca_x_train = pca.fit_transform(self.__x_train.read(key))
            pca_x_test = pca.fit_transform(self.__x_test.read(key))
            self.__x_train.write(key, pca_x_train)
            self.__x_test.write(key, pca_x_test)

    # t-distributed stochastic neighbor embedding
    def TSNE(self, components):
        tsne = TSNE(n_components=components)
        for key in self.__keys:
            tsne_x_train = tsne.fit_transform(self.__x_train.read(key))
            tsne_x_test = tsne.fit_transform(self.__x_test.read(key))
            self.__x_train.write(key, tsne_x_train)
            self.__x_test.write(key, tsne_x_test)

    # set XGBoost model parameter
    def xgboost_set(self, param):
        self.__xgboost.write_all(param)

    # set lightGBM model parameter
    def lightgbm_set(self, param):
        self.__lightgbm.write_all(param)

    def logistic_set(self, param):
        self.__logistic.write_all(param)

    # set SVR model parameter
    def svr_set(self, param):
        self.__svr.write_all(param)

    def grid_search(self, model, param, standard_scaler=False, encode=False):
        for key in self.__keys:
            x_train, x_test, y_train, y_test = self.__read_train_and_test(key)
            if standard_scaler:
                x_train, x_test = self.__standard_scaler(x_train, x_test)
            if encode:
                encode = LabelEncoder()
                y_train = encode.fit_transform(y_train)
            scorer = metrics.make_scorer(metrics.r2_score)
            grid_search = GridSearchCV(estimator=model,
                                       param_grid=param,
                                       scoring=scorer)
            grid_search.fit(x_train, y_train)
            print('----------------------', key, '----------------------')
            print('feature number = ', self.__feature_num.read(key))
            print('best parameters: ', grid_search.best_params_)
            if encode:
                predict = grid_search.predict(x_test)
                predict = encode.inverse_transform(predict)
                print('r2 score: ', metrics.r2_score(y_test, predict))
            else:
                print('r2 score: ', grid_search.score(x_test, y_test))

    def random_search(self, model, param, standard_scaler=False, encode=False):
        for key in self.__keys:
            x_train, x_test, y_train, y_test = self.__read_train_and_test(key)
            if standard_scaler:
                x_train, x_test = self.__standard_scaler(x_train, x_test)
            if encode:
                encode = LabelEncoder()
                y_train = encode.fit_transform(y_train)
            scorer = metrics.make_scorer(metrics.r2_score)
            random_search = RandomizedSearchCV(estimator=model,
                                               param_distributions=param,
                                               scoring=scorer)
            random_search.fit(x_train, y_train)
            print('----------------------', key, '----------------------')
            print('feature number = ', self.__feature_num.read(key))
            print('best parameters: ', random_search.best_params_)
            if encode:
                predict = random_search.predict(x_test)
                predict = encode.inverse_transform(predict)
                print('r2 score: ', metrics.r2_score(y_test, predict))
            else:
                print('r2 score: ', random_search.score(x_test, y_test))

    def xgboost(self, xlsx):
        wb = openpyxl.Workbook()
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            n_estimators = self.__xgboost.read('n_estimators')
            learning_rate = self.__xgboost.read('learning_rate')
            max_depth = self.__xgboost.read('max_depth')

            model = xgb.XGBRegressor(n_estimators=n_estimators,
                                     learning_rate=learning_rate,
                                     max_depth=max_depth)
            # train XGBoost model
            model.fit(x_train, y_train)
            # predict y_test via XGBoost model
            predict = model.predict(x_test)

            # create new sheet
            sheet = wb.create_sheet(key)
            # store data into excel
            self.__store_result_data(key, sheet, predict)
            self.__store_model_set(sheet, self.__xgboost)
            # save the excel
            wb.save(self.__output + xlsx)

            # save XGBoost model
            model_name = self.__output + key + '_xgboost.pickle.dat'
            pickle.dump(model, open(model_name, 'wb'))

    def lightgbm(self, xlsx):
        wb = openpyxl.Workbook()
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            boosting_type = self.__lightgbm.read('boosting_type')
            num_leaves = self.__lightgbm.read('num_leaves')
            learning_rate = self.__lightgbm.read('learning_rate')
            max_depth = self.__lightgbm.read('max_depth')

            model = gbm.LGBMRegressor(boosting_type=boosting_type,
                                      num_leaves=num_leaves,
                                      learning_rate=learning_rate,
                                      max_depth=max_depth)
            # train lightGBM model
            model.fit(x_train, y_train)
            # predict y_test via lightGBM model
            predict = model.predict(x_test)

            # create new sheet
            sheet = wb.create_sheet(key)
            # store data into excel
            self.__store_result_data(key, sheet, predict)
            self.__store_model_set(sheet, self.__lightgbm)
            # save the excel
            wb.save(self.__output + xlsx)

            # save lightGBM model
            model_name = self.__output + key + '_lightgbm.pickle.dat'
            pickle.dump(model, open(model_name, 'wb'))

    def linear_regression(self, xlsx):
        wb = openpyxl.Workbook()
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            x_train, x_test = self.__standard_scaler(x_train, x_test)

            model = LinearRegression()
            # train linear regression model
            model.fit(x_train, y_train)
            # predict y_test via linear regression model
            predict = model.predict(x_test)

            # create new sheet
            sheet = wb.create_sheet(key)
            # store data into excel
            self.__store_result_data(key, sheet, predict)
            # save the excel
            wb.save(self.__output + xlsx)

            # save linear regression model
            model_name = self.__output + key + '_linear.pickle.dat'
            pickle.dump(model, open(model_name, 'wb'))

    def logistic_regression(self, xlsx):
        wb = openpyxl.Workbook()
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            x_train, x_test = self.__standard_scaler(x_train, x_test)
            encoder = LabelEncoder()
            y_train = encoder.fit_transform(y_train)
            max_iter = self.__logistic.read('max_iter')
            random_state = self.__logistic.read('random_state')

            model = LogisticRegression(max_iter=max_iter,
                                       random_state=random_state)
            # train logistic regression model
            model.fit(x_train, y_train)
            # predict y_test via logistic regression model
            predict = model.predict(x_test)
            predict = encoder.inverse_transform(predict)

            # create new sheet
            sheet = wb.create_sheet(key)
            # store data into excel
            self.__store_result_data(key, sheet, predict)
            self.__store_model_set(sheet, self.__logistic)
            # save the excel
            wb.save(self.__output + xlsx)

            # save logistic regression model
            model_name = self.__output + key + '_logistic.pickle.dat'
            pickle.dump(model, open(model_name, 'wb'))

    def svr(self, xlsx):
        wb = openpyxl.Workbook()
        for key in self.__keys:
            x_train, x_test, y_train, _ = self.__read_train_and_test(key)
            x_train, x_test = self.__standard_scaler(x_train, x_test)

            model = svm.SVR(C=self.__svr.read('C'),
                            kernel=self.__svr.read('kernel'),
                            gamma=self.__svr.read('gamma'))
            # train SVR model
            model.fit(x_train, y_train)
            # predict y_test via SVR model
            predict = model.predict(x_test)

            # create new sheet
            sheet = wb.create_sheet(key)
            # store data into excel
            self.__store_result_data(key, sheet, predict)
            self.__store_model_set(sheet, self.__svr)
            # save the excel
            wb.save(self.__output + xlsx)

            # save SVR model
            model_name = self.__output + key + '_svr.pickle.dat'
            pickle.dump(model, open(model_name, 'wb'))

    def display_all_data(self):
        for key in self.__keys:
            print('-----------', key, '-----------')
            print('x_train = ', self.__x_train.read(key))
            print('x_test = ', self.__x_test.read(key))
            print('y_train = ', self.__y_train.read(key))
            print('y_test = ', self.__y_test.read(key))


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dst',
                        default='./result/melt-model/',
                        help='destination path')
    args = parser.parse_args()

    if not os.path.isdir(args.dst):
        os.makedirs(args.dst)

    # train tensile model
    retain_feature = 10
    layer = 70
    tensile_data_set = dataset(const.tensile_key, output=args.dst)
    tensile_data_set.load_data(const.bone_fp, const.bone_ppfp)
    tensile_data_set.reshape_and_repeat((-1, 13), repeat=layer)
    tensile_data_set.mutual_information(retain_feature)
    tensile_data_set.xgboost(xlsx='tensile_xgboost.xlsx')
    tensile_data_set.lightgbm(xlsx='tensile_lightgbm.xlsx')
    tensile_data_set.linear_regression(xlsx='tensile_linear.xlsx')
    tensile_data_set.logistic_regression(xlsx='tensile_logistic.xlsx')
    tensile_data_set.svr(xlsx='tensile_svr.xlsx')

    # train permeability model
    pmb_data_set = dataset(const.pmb_key, output=args.dst)
    pmb_data_set.load_data(const.ring_fp, const.ring_ppfp)
    pmb_data_set.reshape_and_repeat((-1, 13), repeat=layer)
    pmb_data_set.mutual_information(retain_feature)
    pmb_data_set.xgboost(xlsx='pmb_xgboost.xlsx')
    pmb_data_set.lightgbm(xlsx='pmb_lightgbm.xlsx')
    pmb_data_set.linear_regression(xlsx='pmb_linear.xlsx')
    pmb_data_set.logistic_regression(xlsx='pmb_logistic.xlsx')
    pmb_data_set.svr(xlsx='pmb_svr.xlsx')

    # train iron loss model
    iron_data_set = dataset(const.iron_key, output=args.dst)
    iron_data_set.load_data(const.ring_fp, const.ring_ppfp)
    iron_data_set.reshape_and_repeat((-1, 13), repeat=layer)
    iron_data_set.mutual_information(retain_feature)
    iron_data_set.xgboost(xlsx='iron_xgboost.xlsx')
    iron_data_set.lightgbm(xlsx='iron_lightgbm.xlsx')
    iron_data_set.linear_regression(xlsx='iron_linear.xlsx')
    iron_data_set.logistic_regression(xlsx='iron_logistic.xlsx')
    iron_data_set.svr(xlsx='iron_svr.xlsx')
