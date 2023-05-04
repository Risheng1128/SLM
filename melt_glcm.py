import argparse
import cv2 as cv
import numpy as np
import melt_constants as const
import os
import openpyxl
import glob

class GLCM:
    '''
    Parameters
    ----------
    img: array_like, shape=(h,w), dtype=np.uint32
        input image
    vmin: int
        minimum value of input image
    vmax: int
        maximum value of input image
    levels: int
        number of grey-levels of GLCM
    distance: float
        pixel pair distance offsets [pixel] (1.0, 2.0, and etc.)
    angle: float
        pixel pair angles [degree] (0.0, 30.0, 45.0, 90.0, and etc.)
    '''
    def __init__(self, img, vmin=0, vmax=255, levels=8,
                 distance=1.0, angle=0.0):
        self.__img = img
        self.__vmin = vmin
        self.__vmax = vmax
        self.__levels = levels
        self.__distance = distance
        self.__angle = angle
        self.__glcm = self.__get_glcm()
        # may used to compute other features
        self.__feature = {feature: None for feature
                          in ['mean_x', 'mean_y', 'variance_x', 'variance_y']}

    # compute gray-level co-occurrence matrix
    def __get_glcm(self):
        h, w = self.__img.shape
        # digitize
        bins = np.linspace(self.__vmin, self.__vmax + 1, self.__levels + 1)
        gl1 = np.digitize(img, bins) - 1

        # make shifted image
        dx = self.__distance * np.cos(np.deg2rad(self.__angle))
        dy = self.__distance * np.sin(np.deg2rad(-self.__angle))
        mat = np.array([[1.0, 0.0, -dx], [0.0, 1.0, -dy]], dtype=np.float32)
        gl2 = cv.warpAffine(gl1, mat, (w, h), flags=cv.INTER_NEAREST,
                            borderMode=cv.BORDER_REPLICATE)

        # make glcm
        glcm = np.zeros((self.__levels, self.__levels), dtype=np.uint32)

        for i in range(self.__levels):
            for j in range(self.__levels):
                mask = ((gl1 == i) & (gl2 == j))
                glcm[i, j] = mask.sum()

        # eliminate the affect of background
        glcm[:, 0] = 0
        glcm[0, :] = 0
        # normalize
        return glcm / glcm.sum()

    # compute joint mean feature
    def compute_mean(self):
        mean_x = 0
        mean_y = 0

        for i in range(self.__levels):
            for j in range(self.__levels):
                mean_x += self.__glcm[i, j] * i
                mean_y += self.__glcm[i, j] * j

        self.__feature['mean_x'] = mean_x
        self.__feature['mean_y'] = mean_y
        return mean_x, mean_y

    # compute variance feature
    def compute_variance(self):
        if self.__feature['mean_x'] is None or \
           self.__feature['mean_y'] is None:
            self.compute_mean()

        variance_x = 0
        variance_y = 0

        for i in range(self.__levels):
            for j in range(self.__levels):
                variance_x += \
                    self.__glcm[i, j] * ((i - self.__feature['mean_x']) ** 2)
                variance_y += \
                    self.__glcm[i, j] * ((i - self.__feature['mean_y']) ** 2)

        self.__feature['variance_x'] = variance_x
        self.__feature['variance_y'] = variance_y
        return variance_x, variance_y

    # compute standard deviation feature
    def compute_standard_deviation(self):
        if self.__feature['variance_x'] is None or \
           self.__feature['variance_y'] is None:
            self.compute_variance()
        return np.sqrt(variance_x), np.sqrt(variance_y)

    # compute cluster prominence feature
    def compute_cluster_prominence(self):
        if self.__feature['mean_x'] is None or \
           self.__feature['mean_y'] is None:
            self.compute_mean()

        prominence = 0
        mean_x = self.__feature['mean_x']
        mean_y = self.__feature['mean_y']

        for i in range(self.__levels):
            for j in range(self.__levels):
                prominence += \
                    ((i + j - mean_x - mean_y) ** 4) * self.__glcm[i, j]
        return prominence

    # compute cluster shade feature
    def compute_cluster_shade(self):
        if self.__feature['mean_x'] is None or \
           self.__feature['mean_y'] is None:
            self.compute_mean()

        shade = 0
        mean_x = self.__feature['mean_x']
        mean_y = self.__feature['mean_y']

        for i in range(self.__levels):
            for j in range(self.__levels):
                shade += ((i + j - mean_x - mean_y) ** 3) * self.__glcm[i, j]
        return shade

    # compute cluster tendency feature
    def compute_cluster_tendency(self):
        if self.__feature['mean_x'] is None or \
           self.__feature['mean_y'] is None:
            self.compute_mean()

        tendency = 0
        mean_x = self.__feature['mean_x']
        mean_y = self.__feature['mean_y']

        for i in range(self.__levels):
            for j in range(self.__levels):
                tendency += \
                    ((i + j - mean_x - mean_y) ** 2) * self.__glcm[i, j]
        return tendency

    # compute autocorrelation feature
    def compute_autocorrelation(self):
        autocorrelation = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                autocorrelation += self.__glcm[i, j] * i * j
        return autocorrelation

    # compute correlation feature
    def compute_correlation(self):
        if self.__feature['variance_x'] is None or \
           self.__feature['variance_y'] is None:
            self.compute_variance()

        correlation = 0
        mean_x = self.__feature['mean_x']
        mean_y = self.__feature['mean_y']
        variance_x = self.__feature['variance_x']
        variance_y = self.__feature['variance_y']

        for i in range(self.__levels):
            for j in range(self.__levels):
                tmp = self.__glcm[i, j] * (i - mean_x) * (j - mean_y)
                correlation += tmp / np.sqrt(variance_x * variance_y)
        return correlation

    # compute dissimilarity feature
    def compute_dissimilarity(self):
        dissimilarity = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                dissimilarity += self.__glcm[i, j] * np.abs(i - j)
        return dissimilarity

    # compute energy feature
    def compute_energy(self):
        energy = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                energy += self.__glcm[i, j] ** 2
        return energy

    # compute entropy feature
    def compute_entropy(self):
        entropy = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                if self.__glcm[i, j]:
                    entropy += -np.log(self.__glcm[i, j]) * self.__glcm[i, j]
        return entropy

    # compute contrast feature
    def compute_contrast(self):
        contrast = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                contrast += self.__glcm[i, j] * ((i - j) ** 2)
        return contrast

    # compute inverse differential moment (IDM) feature
    def compute_idm(self):
        idm = 0
        for i in range(self.__levels):
            for j in range(self.__levels):
                idm += self.__glcm[i, j] / (1 + (i - j) ** 2)
        return idm

# store data into excel (row direction)
def store_row_data(sheet, datas, base_row, base_col):
    for col, data in zip(range(base_col, base_col + len(datas)), datas):
        sheet.cell(base_row, col).value = data


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--src',
                        default='./data/ct-example/',
                        help='source image path')
    parser.add_argument('--xlsx',
                        default='glcm.xlsx',
                        help='xlsx filename')
    args = parser.parse_args()

    if not args.src:
        raise AssertionError('Source path not found!!')

    item_list = os.listdir(args.src)
    item_list = sorted([i for i in item_list])
    item_list = [args.src + i + '/' for i in item_list]
    item_num = len(item_list)

    wb = openpyxl.Workbook()
    for i in range(item_num):
        layer_list = list(
            enumerate(glob.glob(os.path.join(item_list[i], "*.jpg"))))
        layer_list = sorted([i[1] for i in layer_list])
        layer_num = len(layer_list)

        sheet = wb.create_sheet('item' + str(i))
        # store label
        store_row_data(sheet, const.layer_label + const.feature, 1, 1)

        for j in range(layer_num):
            img = cv.imread(layer_list[j])
            img = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
            feature = {feature: 0 for feature in const.feature}

            for z in [0, 45, 90, 135]:
                glcm = GLCM(img, distance=1, angle=z, levels=8)

                mean_x, mean_y = glcm.compute_mean()
                feature['mean_x'] += mean_x
                feature['mean_y'] += mean_y

                variance_x, variance_y = glcm.compute_variance()
                feature['variance_x'] += variance_x
                feature['variance_y'] += variance_y

                standard_deviation_x, standard_deviation_y = \
                    glcm.compute_standard_deviation()
                feature['standard_deviation_x'] += standard_deviation_x
                feature['standard_deviation_y'] += standard_deviation_y

                feature['contrast'] += glcm.compute_contrast()
                feature['autocorrelation'] += glcm.compute_autocorrelation()
                feature['correlation'] += glcm.compute_correlation()
                feature['dissimilarity'] += glcm.compute_dissimilarity()
                feature['energy'] += glcm.compute_energy()
                feature['entropy'] += glcm.compute_entropy()
                feature['entropy'] += glcm.compute_idm()

            for key, value in feature.items():
                feature[key] = value / 4

            # store feature into excel
            store_row_data(sheet, [j + 1], j + 2, 1)
            datas = [feature[key] for key in feature.keys()]
            store_row_data(sheet, datas, j + 2, 2)
            wb.save(args.xlsx)
            print('finish layer: ', j + 1)

        print('finish item: ', i + 1)
