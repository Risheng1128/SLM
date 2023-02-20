import xgboost as xgb
import pandas as pd
import numpy as np
import openpyxl
import argparse
import pickle
import os
import melt_constants as const

from sklearn import metrics

# load data in excel
def load_data(filepath, keys, header):
    # default add 'layer' header
    remove_header = const.layer_header + header

    x_train = {key: [] for key in keys}
    x_test = {key: [] for key in keys}
    y_train = {key: [] for key in keys}
    y_test = {key: [] for key in keys}

    for file in filepath:
        sheets = pd.read_excel(file, sheet_name=None)

        for sheet in sheets:
            all_data = pd.read_excel(file, sheet_name=sheet)

            for attr, key in zip(header, keys):
                if all_data[attr][0] == const.error_data:
                    continue

                if sheet[-1] == '5' or sheet[-1] == '6':
                    x_test[key].append(all_data.drop(remove_header, axis=1))
                    y_test[key].append(all_data[attr][0])
                else:
                    x_train[key].append(all_data.drop(remove_header, axis=1))
                    y_train[key].append(all_data[attr][0])

    return x_train, x_test, y_train, y_test

def train_xgboost_model(x_train, x_test, y_train, y_test, x_shape,
                        y_repeat, keys, output, xlsx):
    wb = openpyxl.Workbook()

    for key in keys:
        # convert 3-D matrix to 2-D matrix
        x_train[key] = np.reshape(np.array(x_train[key]), x_shape)
        x_test[key] = np.reshape(np.array(x_test[key]), x_shape)

        y_train[key] = np.repeat(np.array(y_train[key]), y_repeat)
        y_test[key] = np.repeat(np.array(y_test[key]), y_repeat)

        model = xgb.XGBRegressor(
            n_estimators=100000, learning_rate=0.1, max_depth=20)
        # train XGBoost model
        model.fit(x_train[key], y_train[key])
        # save XGBoost model
        pickle.dump(model, open(output + key + '.pickle.dat', 'wb'))
        # predict x_test via XGBoost model
        predict = model.predict(x_test[key])

        # create new sheet
        sheet = wb.create_sheet(key)
        for col, header in zip(range(1, len(const.output_header) + 1),
                               const.output_header):
            sheet.cell(1, col).value = header
        sheet.cell(2, 4).value = x_train[key].shape[0]
        sheet.cell(2, 5).value = x_test[key].shape[0]
        # compute R2 score, MSE and MAE
        sheet.cell(2, 6).value = model.score(x_test[key], y_test[key])
        sheet.cell(2, 7).value = \
            metrics.mean_squared_error(y_test[key], predict)
        sheet.cell(2, 8).value = \
            metrics.mean_absolute_error(y_test[key], predict)

        row = 2
        for pre, true in zip(predict, y_test[key]):
            sheet.cell(row, 1).value = pre
            sheet.cell(row, 2).value = true
            sheet.cell(row, 3).value = (abs(pre - true) / true) * 100
            row += 1
        wb.save(output + xlsx)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--dst',
                        default='./result/xgboost/',
                        help='destination path')
    args = parser.parse_args()

    if not os.path.isdir(args.dst):
        os.makedirs(args.dst)

    bone_filepath = ['./data/glcm-data/第一批狗骨頭.xlsx',
                     './data/glcm-data/第二批狗骨頭.xlsx']

    ring_filepath = ['./data/glcm-data/第一批圓環.xlsx',
                     './data/glcm-data/第二批圓環.xlsx']

    x_train, x_test, y_train, y_test = load_data(bone_filepath,
                                                 const.tensile_key,
                                                 const.tensile_header)
    # train tensile model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        x_shape=(-1, 70 * 13), y_repeat=1,
                        keys=const.tensile_key, output=args.dst,
                        xlsx='tensile.xlsx')

    x_train, x_test, y_train, y_test = load_data(ring_filepath,
                                                 const.magnetic_key +
                                                 const.iron_key,
                                                 const.magnetic_header +
                                                 const.iron_header)
    # train permeability model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        x_shape=(-1, 13), y_repeat=70,
                        keys=const.magnetic_key, output=args.dst,
                        xlsx='magnetic.xlsx')
    # train iron loss model
    train_xgboost_model(x_train, x_test, y_train, y_test,
                        x_shape=(-1, 70 * 13), y_repeat=1,
                        keys=const.iron_key, output=args.dst,
                        xlsx='iron.xlsx')
